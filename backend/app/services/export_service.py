from datetime import datetime
from pathlib import Path
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.repositories.invoice_repository import InvoiceRepository
from app.schemas.export import ExportFilter


class ExportService:

    def __init__(self, db):
        self.repository = InvoiceRepository(db)

    def export_excel(
        self,
        user_id: UUID,
        filters: ExportFilter,
    ) -> str:

        invoices = self.repository.export_filtered(
            user_id,
            filters,
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Invoice Report"

        # =====================================================
        # Colors
        # =====================================================

        blue_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78",
        )

        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE",
            end_color="C6EFCE",
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="FFC7CE",
            end_color="FFC7CE",
        )

        gray_fill = PatternFill(
            fill_type="solid",
            start_color="F5F5F5",
            end_color="F5F5F5",
        )

        summary_fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3",
        )

        white_font = Font(
            bold=True,
            color="FFFFFF",
        )

        bold_font = Font(
            bold=True,
        )

        title_font = Font(
            bold=True,
            size=18,
            color="1F4E78",
        )

        subtitle_font = Font(
            italic=True,
            size=11,
        )

        thin = Side(
            style="thin",
            color="D9D9D9",
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        center = Alignment(
            horizontal="center",
            vertical="center",
        )

        # =====================================================
        # Report Title
        # =====================================================

        worksheet.merge_cells("A1:H1")
        worksheet["A1"] = "AI BILL PROCESSING SYSTEM"
        worksheet["A1"].font = title_font
        worksheet["A1"].alignment = center

        worksheet.merge_cells("A2:H2")
        worksheet["A2"] = "Invoice Export Report"
        worksheet["A2"].font = bold_font
        worksheet["A2"].alignment = center

        worksheet.merge_cells("A3:H3")
        worksheet["A3"] = (
            "Generated On : "
            + datetime.now().strftime(
                "%d-%b-%Y %I:%M %p"
            )
        )
        worksheet["A3"].font = subtitle_font
        worksheet["A3"].alignment = center

        # =====================================================
        # Summary
        # =====================================================

        total_invoices = len(invoices)

        verified = len(
            [
                invoice
                for invoice in invoices
                if invoice.gst_match
            ]
        )

        mismatched = len(
            [
                invoice
                for invoice in invoices
                if invoice.gst_match is False
            ]
        )

        total_amount = sum(
            float(invoice.total_amount or 0)
            for invoice in invoices
        )

        worksheet["A5"] = "Total Invoices"
        worksheet["B5"] = total_invoices

        worksheet["D5"] = "GST Verified"
        worksheet["E5"] = verified

        worksheet["G5"] = "GST Mismatch"
        worksheet["H5"] = mismatched

        worksheet["A6"] = "Total Amount"
        worksheet["B6"] = total_amount

        for cell in [
            "A5",
            "D5",
            "G5",
            "A6",
        ]:
            worksheet[cell].font = bold_font
            worksheet[cell].fill = summary_fill

        worksheet["B6"].number_format = "₹#,##0.00"

        # =====================================================
        # Table starts on Row 8
        # =====================================================

        row = 8
        # =====================================================
        # Table Headers
        # =====================================================

        headers = [
            "Invoice Number",
            "Firm Name",
            "Vendor",
            "GST Number",
            "Invoice Date",
            "Total Amount",
            "GST Status",
            "Verification Message",
        ]

        for column, header in enumerate(headers, start=1):

            cell = worksheet.cell(
                row=row,
                column=column,
            )

            cell.value = header
            cell.font = white_font
            cell.fill = blue_fill
            cell.border = border
            cell.alignment = center

        row += 1

        # =====================================================
        # Invoice Data
        # =====================================================

        for index, invoice in enumerate(invoices):

            # Zebra Striping

            current_fill = (
                gray_fill
                if index % 2 == 0
                else PatternFill(fill_type=None)
            )

            worksheet.cell(
                row=row,
                column=1,
            ).value = invoice.invoice_number or "-"

            worksheet.cell(
                row=row,
                column=2,
            ).value = (
                invoice.firm.name
                if invoice.firm
                else "-"
            )

            worksheet.cell(
                row=row,
                column=3,
            ).value = invoice.vendor_name or "-"

            worksheet.cell(
                row=row,
                column=4,
            ).value = invoice.gst_number or "-"

            # Date

            date_cell = worksheet.cell(
                row=row,
                column=5,
            )

            if invoice.invoice_date:

                date_cell.value = invoice.invoice_date
                date_cell.number_format = "DD-MMM-YYYY"

            else:

                date_cell.value = "-"

            # Amount

            amount_cell = worksheet.cell(
                row=row,
                column=6,
            )

            amount_cell.value = (
                float(invoice.total_amount)
                if invoice.total_amount
                else 0
            )

            amount_cell.number_format = "₹#,##0.00"

            # GST Status

            gst_cell = worksheet.cell(
                row=row,
                column=7,
            )

            if invoice.gst_match is True:

                gst_cell.value = "VERIFIED"
                gst_cell.fill = green_fill

            elif invoice.gst_match is False:

                gst_cell.value = "MISMATCH"
                gst_cell.fill = red_fill

            else:

                gst_cell.value = "PENDING"

            worksheet.cell(
                row=row,
                column=8,
            ).value = (
                invoice.verification_message
                or "-"
            )

            # Apply Styling

            for col in range(1, 9):

                cell = worksheet.cell(
                    row=row,
                    column=col,
                )

                cell.border = border

                if (
                    col != 7
                    and current_fill.fill_type
                ):
                    cell.fill = current_fill

                if col in [5, 6, 7]:
                    cell.alignment = center

            row += 1
        # =====================================================
        # Auto Adjust Column Width
        # =====================================================

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value)),
                        )

                except Exception:
                    pass

            adjusted_width = min(
                max_length + 4,
                40,
            )

            worksheet.column_dimensions[
                column_letter
            ].width = adjusted_width

        # =====================================================
        # Freeze Header Row
        # =====================================================

        worksheet.freeze_panes = "A9"

        # =====================================================
        # Enable Auto Filter
        # =====================================================

        worksheet.auto_filter.ref = (
            f"A8:H{row-1}"
        )

        # =====================================================
        # Page Setup
        # =====================================================

        worksheet.sheet_view.showGridLines = False

        worksheet.page_setup.orientation = (
            "landscape"
        )

        worksheet.page_setup.fitToWidth = 1

        worksheet.page_margins.left = 0.3
        worksheet.page_margins.right = 0.3
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5

        # =====================================================
        # Footer
        # =====================================================

        footer_row = row + 2

        worksheet.merge_cells(
            start_row=footer_row,
            start_column=1,
            end_row=footer_row,
            end_column=8,
        )

        footer = worksheet.cell(
            row=footer_row,
            column=1,
        )

        footer.value = (
            "Generated automatically by "
            "AI Bill Processing System"
        )

        footer.font = Font(
            italic=True,
            color="808080",
        )

        footer.alignment = center

        # =====================================================
        # Save Workbook
        # =====================================================

        export_dir = Path("exports")

        export_dir.mkdir(
            exist_ok=True,
        )

        filename = (
            "Invoice_Report_"
            + datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )
            + ".xlsx"
        )

        file_path = export_dir / filename

        workbook.save(file_path)

        return str(file_path)
